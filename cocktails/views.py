from django.db.models import ProtectedError
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Cocktail, UsersFavouriteCocktails, Ingredient, CocktailIngrediets
from .forms import CocktailForm, CocktailIngredientFormSet, IngredientForm
from django.core.paginator import Paginator
import openpyxl
from django.http import HttpResponse
from django.shortcuts import redirect


def cocktails_list(request):
    cocktails = Cocktail.objects.all().order_by('-id')
    paginator = Paginator(cocktails, 5)
    page = request.GET.get('page')
    cocktails = paginator.get_page(page)
    return render(request, 'cocktails/cocktails_list.html', {'cocktails':cocktails})

def user_cocktails(request):
    cocktails = request.user.cocktails.all()
    return render(request, 'cocktails/user_cocktails.html', {'user': request.user, 'cocktails':cocktails})

@login_required
def create_cocktail(request):
    if request.method == 'POST':
        form = CocktailForm(request.POST, request.FILES)
        formset = CocktailIngredientFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            cocktail = form.save(commit=False)
            cocktail.user = request.user
            cocktail.save()
            formset.instance = cocktail
            formset.save()
            return redirect('/cocktails/')
    else:
        form = CocktailForm()
        formset = CocktailIngredientFormSet()

    return render(request, 'cocktails/create_cocktail.html', {
        'form': form,
        'formset': formset,
    })

@login_required
def edit_cocktail(request, pk):
    cocktail = get_object_or_404(Cocktail, pk=pk)

    if request.method == 'POST':
        form = CocktailForm(request.POST, request.FILES, instance=cocktail)
        formset = CocktailIngredientFormSet(request.POST, instance=cocktail)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect('cocktail_details', pk=pk)
    else:
        form = CocktailForm(instance=cocktail)
        formset = CocktailIngredientFormSet(instance=cocktail)

    return render(request, 'cocktails/edit_cocktail.html', {
        'cocktail': cocktail,
        'form': form,
        'formset': formset,
    })

@login_required
def delete_cocktail(request, pk):
    cocktail = get_object_or_404(Cocktail, pk=pk, user=request.user)
    if request.method == 'POST':
        cocktail.delete()
        return redirect('/cocktails/')

@login_required
def starred_cocktails(request):
    favourites = UsersFavouriteCocktails.objects.filter(user=request.user).values_list('cocktail', flat=True)
    cocktails = Cocktail.objects.filter(id__in=favourites)
    return render(request, 'cocktails/user_favourite_cocktails.html', {'cocktails': cocktails})

@login_required
def add_favourite(request, pk):
    cocktail = get_object_or_404(Cocktail, pk=pk)
    UsersFavouriteCocktails.objects.get_or_create(user=request.user, cocktail=cocktail)
    return redirect('/cocktails/')

@login_required
def remove_favourite(request, pk):
    UsersFavouriteCocktails.objects.filter(user=request.user, cocktail__pk=pk).delete()
    return redirect('/cocktails/')

@login_required
def toggle_star(request, pk):
    cocktail = get_object_or_404(Cocktail, pk=pk)
    fav, created = UsersFavouriteCocktails.objects.get_or_create(
        user=request.user, cocktail=cocktail)
    if not created:
        fav.delete()
    return redirect(request.META.get('HTTP_REFERER', '/cocktails/'))

@login_required
def ingredients_list(request):
    ingredients = Ingredient.objects.all()
    paginator = Paginator(ingredients, 5)
    page = request.GET.get('page')
    ingredients = paginator.get_page(page)
    return render(request, 'ingredients/ingredients_list.html', {'ingredients':ingredients})

from .models import CocktailIngrediets

def cocktail_details(request, pk):
    cocktail = get_object_or_404(Cocktail, pk=pk)
    ingredients = CocktailIngrediets.objects.filter(cocktail=cocktail).select_related('ingredient')
    return render(request, 'cocktails/cocktail_details.html', {
        'cocktail': cocktail,
        'ingredients': ingredients,
    })




@login_required
def edit_ingredient(request, pk):
    ingredient = get_object_or_404(Ingredient, pk=pk)
    if request.method == 'POST':
        form = IngredientForm(request.POST, request.FILES, instance=ingredient)
        if form.is_valid():
            form.save()
            return redirect('ingredients_list')
    else:
        form = IngredientForm(instance=ingredient)

    return render(request, 'ingredients/edit_ingredient.html', {'form': form})

@login_required
def delete_ingredient(request, pk):
    ingredient = get_object_or_404(Ingredient, pk=pk)
    form = IngredientForm(request.POST, request.FILES, instance=ingredient)
    if request.method == 'POST':
        try:
            ingredient.delete()
            return redirect('ingredients_list')
        except ProtectedError:
            return render(request, 'ingredients/delete_ingredient.html', {'ingredient': ingredient,
                'error': 'You cannot delete an ingredient connected to a cocktail'})


@login_required
def create_ingredient(request):
    if request.method == 'POST':
        form = IngredientForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('ingredients_list')
    else:
        form = IngredientForm()
    return render(request, 'ingredients/create_ingredient.html', {'form': form})

def export_cocktails_to_excel(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_cocktails')
        if not selected_ids:
            return redirect('cocktails')

        cocktails = Cocktail.objects.filter(id__in=selected_ids)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chosen cocktails"

        headers = ['Name', 'Category', 'Ingredients']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        for cocktail in cocktails:
            ingredients_qs = cocktail.cocktailingrediets_set.all()

            ingredients_list = []
            for ci in ingredients_qs:
                ingredients_list.append(f"{ci.ingredient.name} ({ci.quantity} {ci.unit})")

            ingredients_string = ", ".join(ingredients_list)
            ws.append([cocktail.name, cocktail.category, ingredients_string])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="your_cocktails.xlsx"'
        wb.save(response)

        return response

    return redirect('cocktails')